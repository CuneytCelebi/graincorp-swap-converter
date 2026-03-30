import streamlit as st
import json, base64, io, csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import urllib.request, urllib.error
import pandas as pd

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GrainCorp Swap Converter",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.banner {
    background: linear-gradient(135deg, #1B5E3B 0%, #2E7D52 100%);
    color: white; padding: 24px 32px; border-radius: 12px; margin-bottom: 24px;
}
.banner h1 { margin: 0; font-size: 26px; font-weight: 700; }
.banner p  { margin: 4px 0 0; opacity: 0.75; font-size: 14px; }
.card { background: white; border: 1px solid #E2E6EA; border-radius: 10px;
        padding: 20px 24px; margin-bottom: 16px; }
.sec-title { font-size: 12px; font-weight: 700; color: #6B7C72;
             text-transform: uppercase; letter-spacing: .08em; margin-bottom: 10px; }
.badge-ok      { background:#E8F5EE; color:#1B7A3E; padding:3px 10px;
                 border-radius:20px; font-size:12px; font-weight:600; }
.badge-warn    { background:#FEF8EC; color:#B7600A; padding:3px 10px;
                 border-radius:20px; font-size:12px; font-weight:600; }
.badge-error   { background:#FDEDEC; color:#C0392B; padding:3px 10px;
                 border-radius:20px; font-size:12px; font-weight:600; }
.badge-pending { background:#EFF0F1; color:#6B7C72; padding:3px 10px;
                 border-radius:20px; font-size:12px; font-weight:600; }
.mismatch-row { background:#FEF8EC; border-left:4px solid #E8971A;
                padding:10px 14px; border-radius:4px; margin-bottom:6px;
                font-size:13px; }
.match-row    { background:#E8F5EE; border-left:4px solid #1B7A3E;
                padding:10px 14px; border-radius:4px; margin-bottom:6px;
                font-size:13px; }
.hist-table { width:100%; border-collapse:collapse; font-size:13px; }
.hist-table th { background:#1B5E3B; color:white; padding:10px 12px;
                 text-align:left; font-weight:600; font-size:12px; }
.hist-table td { padding:9px 12px; border-bottom:1px solid #F0F1F2; color:#2C3E30; }
.hist-table tr:nth-child(even) td { background:#FAFBFC; }
.metric-card { background:white; border:1px solid #E2E6EA; border-radius:10px;
               padding:16px 20px; text-align:center; }
.metric-val  { font-size:28px; font-weight:700; color:#1B5E3B; }
.metric-lbl  { font-size:12px; color:#9DAAA2; margin-top:4px; }
.stButton > button { border-radius:8px !important; font-weight:600 !important;
                     border:none !important; }
.stProgress > div > div { background:#1B5E3B !important; }
</style>
""", unsafe_allow_html=True)

# ── Config ─────────────────────────────────────────────────────────────────
HEADERS_ROW = [
    'Swap No','Document Date','GHA Site Code (FROM)','GHA Site Code (To)',
    'GHA','Site From','Site To','Rate From','Rate To','Grade ',
    'Season','Tonnes','Value ex GST'
]

EXTRACT_PROMPT = """Extract data from this GrainCorp Customer Stock Swap PDF.
Return ONLY a JSON object with no markdown formatting:
{
  "swap_no": "",
  "transfer_date": "",
  "from_site_code": "",
  "to_site_code": "",
  "from_site_name": "",
  "to_site_name": "",
  "rate_from": null,
  "rate_to": null,
  "grade": "",
  "season": null,
  "tonnes": null
}
Rules:
- swap_no: number only e.g. "6001086404"
- transfer_date: DD/MM/YYYY e.g. "23/01/2026"
- from_site_code: numeric code from FROM table e.g. "5168"
- to_site_code: numeric code from TO table e.g. "4207"
- from_site_name: "2203 SITENAME GC" e.g. "2203 MIRROOL GC"
- to_site_name: "2203 SITENAME GC" e.g. "2203 RED BEND GC"
- rate_from: freight $ FROM side as number e.g. 51.20
- rate_to: freight $ TO side as number e.g. 56.52
- grade: e.g. "H2" or "APW1"
- season: 4-digit int, slash removed e.g. "25/26"->2526
- tonnes: e.g. 1688.51
Return ONLY the JSON."""

# ── Session state ──────────────────────────────────────────────────────────
for k, v in [("history",[]), ("log",[]), ("rate_table", {})]:
    if k not in st.session_state:
        st.session_state[k] = v

def add_log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state.log.insert(0, {"ts":ts, "msg":msg, "level":level})
    st.session_state.log = st.session_state.log[:200]

# ── Rate table helpers ─────────────────────────────────────────────────────
def load_rate_table_from_df(df):
    """Parse uploaded dataframe into {site_code: {name, rate_from, rate_to}}"""
    df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
    rate_map = {}
    for _, row in df.iterrows():
        code = str(row.get("site_code", row.get("code",""))).strip()
        if not code or code == "nan": continue
        rate_map[code] = {
            "site_name":  str(row.get("site_name", row.get("name",""))).strip(),
            "rate_from":  _safe_float(row.get("rate_from", row.get("from_rate",
                          row.get("freight_from", None)))),
            "rate_to":    _safe_float(row.get("rate_to",   row.get("to_rate",
                          row.get("freight_to", None)))),
        }
    return rate_map

def _safe_float(val):
    try: return float(val) if val is not None and str(val) != "nan" else None
    except: return None

def check_rates(data, rate_table):
    """Returns list of mismatch dicts for a single extracted record."""
    mismatches = []
    if not rate_table: return mismatches

    from_code = str(data.get("from_site_code","")).strip()
    to_code   = str(data.get("to_site_code","")).strip()

    if from_code in rate_table:
        expected = rate_table[from_code].get("rate_from")
        actual   = _safe_float(data.get("rate_from"))
        if expected is not None and actual is not None and abs(expected - actual) > 0.01:
            mismatches.append({
                "field":    "Rate From",
                "site":     data.get("from_site_name", from_code),
                "expected": expected,
                "actual":   actual,
                "diff":     actual - expected,
            })

    if to_code in rate_table:
        expected = rate_table[to_code].get("rate_to")
        actual   = _safe_float(data.get("rate_to"))
        if expected is not None and actual is not None and abs(expected - actual) > 0.01:
            mismatches.append({
                "field":    "Rate To",
                "site":     data.get("to_site_name", to_code),
                "expected": expected,
                "actual":   actual,
                "diff":     actual - expected,
            })
    return mismatches

# ── Gemini helpers ─────────────────────────────────────────────────────────
def get_best_model(api_key):
    preferred = ["gemini-2.5-flash","gemini-2.0-flash","gemini-2.0-flash-lite",
                 "gemini-1.5-flash","gemini-pro"]
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
        with urllib.request.urlopen(url, timeout=10) as r:
            data = json.loads(r.read())
        avail = [m["name"].replace("models/","") for m in data.get("models",[])
                 if "generateContent" in m.get("supportedGenerationMethods",[])]
        for p in preferred:
            if p in avail: return p
        if avail: return avail[0]
    except: pass
    return "gemini-2.5-flash"

def call_gemini(api_key, pdf_bytes):
    b64   = base64.b64encode(pdf_bytes).decode()
    model = get_best_model(api_key)
    payload = json.dumps({
        "contents":[{"parts":[
            {"inline_data":{"mime_type":"application/pdf","data":b64}},
            {"text":EXTRACT_PROMPT}
        ]}],
        "generationConfig":{"temperature":0,"maxOutputTokens":2048}
    }).encode()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    req = urllib.request.Request(url, data=payload,
                                 headers={"Content-Type":"application/json"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read())
    except urllib.error.HTTPError as e:
        raise Exception(f"API error {e.code}: {e.read().decode()[:200]}")
    cand = data.get("candidates",[{}])[0]
    fin  = cand.get("finishReason","STOP")
    if fin not in ("STOP",""):
        raise Exception(f"Gemini stopped early ({fin}). Try again.")
    text = cand["content"]["parts"][0]["text"]
    text = text.replace("```json","").replace("```","").strip()
    s,e  = text.find("{"), text.rfind("}")+1
    if s>=0 and e>s: text=text[s:e]
    return json.loads(text), model

# ── Excel builder ──────────────────────────────────────────────────────────
# Highlight colours
MISMATCH_FILL = PatternFill("solid", start_color="FFE0B2")  # orange
MISMATCH_FONT_COLOR = "BF360C"
NORMAL_FILL   = PatternFill("solid", start_color="FFFFFFCC")

def build_excel(rows, mismatches_by_swap):
    """
    rows: list of extracted dicts
    mismatches_by_swap: {swap_no: [mismatch_dict, ...]}
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "MILLING"
    thin = Side(style="thin")
    def B(): return Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.row_dimensions[1].height = 104.25
    for col, hdr in enumerate(HEADERS_ROW, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font = Font(name="Arial", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = B()

    for ri, r in enumerate(rows, 2):
        swap_no   = str(r.get("swap_no",""))
        mismatches= mismatches_by_swap.get(swap_no, [])
        mismatch_fields = {m["field"] for m in mismatches}

        def sc(col, val, highlight=False):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Calibri", size=11,
                          color=MISMATCH_FONT_COLOR if highlight else "000000",
                          bold=highlight)
            c.fill = MISMATCH_FILL if highlight else NORMAL_FILL
            c.border = B()
            return c

        sc(1, int(r["swap_no"]) if r.get("swap_no") else None)
        dv=None; ds=r.get("transfer_date","")
        if ds:
            try: dv=datetime.strptime(ds,"%d/%m/%Y")
            except: dv=ds
        dc=sc(2,dv); dc.number_format="DD/MM/YYYY"
        sc(3, int(r["from_site_code"]) if r.get("from_site_code") else None)
        sc(4, int(r["to_site_code"])   if r.get("to_site_code")   else None)
        sc(5, "GC")
        sc(6, r.get("from_site_name",""))
        sc(7, r.get("to_site_name",""))

        # Rate From — highlight if mismatch
        rf = sc(8, r.get("rate_from"), highlight="Rate From" in mismatch_fields)
        rf.number_format = "0.00"
        # Rate To — highlight if mismatch
        rt = sc(9, r.get("rate_to"),   highlight="Rate To" in mismatch_fields)
        rt.number_format = "0.00"

        sc(10, r.get("grade",""))
        sc(11, int(r["season"]) if r.get("season") else None)
        tc = sc(12, r.get("tonnes"))
        tc.number_format = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'

        vc = ws.cell(row=ri, column=13)
        vc.value = f"=(H{ri}-I{ri})*L{ri}"
        vc.font  = Font(name="Calibri", size=11)
        vc.fill  = NORMAL_FILL; vc.border = B()
        vc.number_format = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

    for i,w in enumerate([15,11.86,10.57,10.57,8,25,25,10.57,10.57,8,8,18.86,16.29],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F2"

    last_row = len(rows)+1
    if last_row >= 2:
        tbl = Table(displayName="SwapData", ref=f"A1:M{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium16",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)

    # ── Mismatches sheet ──
    total_mismatches = sum(len(v) for v in mismatches_by_swap.values())
    if total_mismatches > 0:
        ws2 = wb.create_sheet("Rate Mismatches")
        ws2.append(["Swap No","Field","Site","Expected Rate","Actual Rate","Difference"])
        hdr_fill = PatternFill("solid", start_color="BF360C")
        for col in range(1, 7):
            c = ws2.cell(row=1, column=col)
            c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.border = B()
        r = 2
        for swap_no, mismatches in mismatches_by_swap.items():
            for m in mismatches:
                ws2.append([swap_no, m["field"], m["site"],
                            m["expected"], m["actual"],
                            round(m["diff"], 4)])
                for col in range(1, 7):
                    c = ws2.cell(row=r, column=col)
                    c.font  = Font(name="Calibri", size=11)
                    c.fill  = PatternFill("solid", start_color="FFF3E0")
                    c.border = B()
                r += 1
        for i, w in enumerate([14,12,28,16,16,14], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def make_rate_template():
    """Generate a downloadable rate table template."""
    buf = io.BytesIO()
    wb  = openpyxl.Workbook(); ws = wb.active; ws.title = "Rates"
    headers = ["site_code","site_name","rate_from","rate_to"]
    thin = Side(style="thin")
    def B(): return Border(left=thin,right=thin,top=thin,bottom=thin)
    hfill = PatternFill("solid", start_color="1B5E3B")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF"); c.fill = hfill; c.border = B()
    samples = [
        ["5168","2203 MIRROOL GC",51.20,""],
        ["4207","2203 RED BEND GC","",56.52],
        ["6104","2203 GILGANDRA GC",59.33,63.59],
    ]
    for ri, row in enumerate(samples, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=11); c.border = B()
    for i, w in enumerate([12, 26, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style='background:linear-gradient(135deg,#1B5E3B,#2E7D52);
                padding:20px;border-radius:10px;margin-bottom:20px;'>
        <div style='color:white;font-size:20px;font-weight:700;'>🌾 GrainCorp</div>
        <div style='color:#A8D5B5;font-size:12px;margin-top:4px;'>Swap Converter</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("### 🔑 API Key")
    api_key = st.text_input("Google Gemini API Key", type="password",
                             placeholder="AIzaSy...",
                             help="Get a free key at aistudio.google.com")
    if api_key:
        st.success("✓ API key entered")
    else:
        st.info("Enter your Gemini API key.\n\n"
                "Get one free at [aistudio.google.com](https://aistudio.google.com/apikey)")

    st.divider()
    rt = st.session_state.rate_table
    total_rates = len(rt)
    total_hist  = len(st.session_state.history)
    st.markdown("### 📊 Stats")
    c1, c2 = st.columns(2)
    c1.metric("Rate Entries", total_rates)
    c2.metric("Extracted", total_hist)

    if total_rates:
        st.success(f"✓ Rate table loaded — {total_rates} sites")
    else:
        st.warning("⚠️ No rate table loaded.\nUpload one in the **Rates** tab.")

    st.divider()
    st.markdown("""<div style='font-size:11px;color:#9DAAA2;line-height:1.8;'>
    <b>Free Tier Limits</b><br>
    Gemini 2.5 Flash: ~20 req/day<br>
    Gemini 2.0 Flash: ~200 req/day<br><br>
    <b>Privacy</b><br>PDFs sent to Google Gemini only.
    Nothing stored on any server.</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════
st.markdown("""<div class='banner'>
    <h1>🌾 GrainCorp Swap Converter</h1>
    <p>PDF → Excel  •  Rate Validation  •  Powered by Google Gemini (Free)</p>
</div>""", unsafe_allow_html=True)

tab_convert, tab_rates, tab_history, tab_log = st.tabs(
    ["⚡  Convert", "📊  Rate Table", "📋  History", "🔍  Log"])


# ══════════════════════════════════════════════════════════════════════════
# RATES TAB
# ══════════════════════════════════════════════════════════════════════════
with tab_rates:
    st.markdown("### 📊 Site Rate Table")
    st.markdown("""
    Upload a CSV or Excel file with your expected freight rates per site.
    These will be compared against the rates in each PDF — any mismatches will be
    **highlighted in orange** in the Excel output and listed on a separate sheet.
    """)

    # Template download
    col1, col2 = st.columns([2,1])
    with col1:
        st.markdown("#### Required columns")
        st.markdown("""
        | Column | Description | Example |
        |---|---|---|
        | `site_code` | Numeric site code | `5168` |
        | `site_name` | Site name | `2203 MIRROOL GC` |
        | `rate_from` | Expected freight rate FROM this site | `51.20` |
        | `rate_to` | Expected freight rate TO this site | `56.52` |

        Leave `rate_from` or `rate_to` blank if a site only applies one way.
        """)
    with col2:
        st.markdown("#### Download template")
        st.download_button(
            "📥 Download Rate Template",
            data=make_rate_template(),
            file_name="GrainCorp_Rate_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.divider()

    # Upload
    st.markdown("#### Upload your rate file")
    rate_file = st.file_uploader("Rate file", type=["csv","xlsx","xls"],
                                  label_visibility="collapsed")

    if rate_file:
        try:
            if rate_file.name.endswith(".csv"):
                df = pd.read_csv(rate_file)
            else:
                df = pd.read_excel(rate_file)

            rate_map = load_rate_table_from_df(df)
            if rate_map:
                st.session_state.rate_table = rate_map
                st.success(f"✅ Loaded {len(rate_map)} site(s) from **{rate_file.name}**")
                add_log(f"Rate table loaded: {len(rate_map)} sites from {rate_file.name}", "ok")
            else:
                st.error("❌ Could not find required columns. "
                         "Make sure your file has: site_code, rate_from, rate_to")
        except Exception as e:
            st.error(f"❌ Error reading file: {e}")

    # Show current rate table
    if st.session_state.rate_table:
        st.divider()
        st.markdown(f"#### Current rate table — {len(st.session_state.rate_table)} sites")

        rows_html = ""
        for code, info in st.session_state.rate_table.items():
            rf = f"${info['rate_from']:.2f}" if info.get('rate_from') is not None else "—"
            rt = f"${info['rate_to']:.2f}"   if info.get('rate_to')   is not None else "—"
            rows_html += f"""<tr>
                <td><code>{code}</code></td>
                <td>{info.get('site_name','—')}</td>
                <td style='color:#1B5E3B;font-weight:600;'>{rf}</td>
                <td style='color:#1565C0;font-weight:600;'>{rt}</td>
            </tr>"""

        st.markdown(f"""<table class='hist-table'>
            <thead><tr>
                <th>Site Code</th><th>Site Name</th>
                <th>Rate From ($)</th><th>Rate To ($)</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
        </table>""", unsafe_allow_html=True)

        if st.button("🗑️ Clear Rate Table"):
            st.session_state.rate_table = {}
            st.rerun()
    else:
        st.info("No rate table loaded yet. Upload a file above or download the template to get started.")


# ══════════════════════════════════════════════════════════════════════════
# CONVERT TAB
# ══════════════════════════════════════════════════════════════════════════
with tab_convert:
    # Stats row
    rate_table = st.session_state.rate_table
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f"""<div class='metric-card'>
        <div class='metric-val'>PDF</div><div class='metric-lbl'>Input Format</div>
    </div>""", unsafe_allow_html=True)
    c2.markdown(f"""<div class='metric-card'>
        <div class='metric-val'>{'✓' if rate_table else '—'}</div>
        <div class='metric-lbl'>Rate Table {'Active' if rate_table else 'Not Loaded'}</div>
    </div>""", unsafe_allow_html=True)
    c3.markdown(f"""<div class='metric-card'>
        <div class='metric-val'>{len(rate_table)}</div>
        <div class='metric-lbl'>Sites in Rate Table</div>
    </div>""", unsafe_allow_html=True)
    c4.markdown(f"""<div class='metric-card'>
        <div class='metric-val'>{len(st.session_state.history)}</div>
        <div class='metric-lbl'>Total Extracted</div>
    </div>""", unsafe_allow_html=True)

    if not rate_table:
        st.warning("⚠️ No rate table loaded — rates won't be validated. "
                   "Upload one in the **Rate Table** tab to enable mismatch highlighting.")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("#### 📄 Upload Swap PDFs")

    uploaded_files = st.file_uploader("PDFs", type=["pdf"],
                                       accept_multiple_files=True,
                                       label_visibility="collapsed")

    if uploaded_files:
        st.markdown(f"**{len(uploaded_files)} file(s) ready:**")
        for f in uploaded_files:
            st.markdown(f"""<div style='display:flex;align-items:center;gap:10px;
                            padding:8px 12px;background:white;border:1px solid #E2E6EA;
                            border-radius:8px;margin-bottom:4px;'>
                <span>📄</span>
                <div style='flex:1;'>
                    <div style='font-weight:600;font-size:14px;'>{f.name}</div>
                    <div style='color:#9DAAA2;font-size:12px;'>{f.size/1024:.1f} KB</div>
                </div>
                <span class='badge-pending'>READY</span>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if not api_key:
            st.warning("⚠️ Please enter your Gemini API key in the sidebar first.")
        else:
            if st.button("⚡  Extract All & Download Excel",
                         type="primary", use_container_width=True):

                results = []
                all_mismatches = {}
                errors = []
                progress = st.progress(0, text="Starting...")
                status   = st.empty()

                for i, f in enumerate(uploaded_files):
                    status.info(f"Processing **{f.name}** ({i+1}/{len(uploaded_files)})...")
                    try:
                        data, model = call_gemini(api_key, f.read())

                        # Rate check
                        mismatches = check_rates(data, rate_table)
                        swap_no    = str(data.get("swap_no",""))
                        if mismatches:
                            all_mismatches[swap_no] = mismatches
                            add_log(f"⚠ Swap {swap_no}: {len(mismatches)} rate mismatch(es)", "warn")
                        else:
                            add_log(f"✓ Swap {swap_no} extracted, rates OK [{model}]", "ok")

                        results.append(data)
                        st.session_state.history.insert(0, {
                            **data,
                            "source_file":   f.name,
                            "extracted_at":  datetime.now().strftime("%d/%m/%Y %H:%M"),
                            "mismatches":    len(mismatches),
                        })
                    except Exception as e:
                        errors.append(f.name)
                        add_log(f"✗ Failed: {f.name} — {e}", "error")

                    progress.progress((i+1)/len(uploaded_files),
                                      text=f"Processed {i+1}/{len(uploaded_files)}")

                progress.empty()

                if results:
                    total_mm = sum(len(v) for v in all_mismatches.values())

                    if total_mm > 0:
                        status.warning(f"⚠️ Extracted {len(results)} record(s) — "
                                       f"**{total_mm} rate mismatch(es) found** "
                                       f"(highlighted in Excel)")
                    else:
                        status.success(f"✅ Extracted {len(results)} record(s) — "
                                       f"all rates match!")

                    if errors:
                        st.warning(f"⚠️ {len(errors)} file(s) failed: {', '.join(errors)}")

                    # ── Mismatch summary ──
                    if total_mm > 0:
                        st.markdown("#### ⚠️ Rate Mismatches Found")
                        st.markdown("The following rates differ from your rate table. "
                                    "They will be **highlighted in orange** in the Excel output "
                                    "and listed on a separate **Rate Mismatches** sheet.")

                        for swap_no, mismatches in all_mismatches.items():
                            for m in mismatches:
                                diff_sign = "+" if m["diff"] > 0 else ""
                                diff_col  = "#C0392B" if m["diff"] > 0 else "#1565C0"
                                st.markdown(f"""
                                <div class='mismatch-row'>
                                    <b>Swap {swap_no}</b> — {m['field']} — {m['site']}<br>
                                    <span style='color:#6B7C72;'>Expected:</span>
                                    <b>${m['expected']:.2f}</b>
                                    &nbsp;→&nbsp;
                                    <span style='color:#6B7C72;'>Actual:</span>
                                    <b>${m['actual']:.2f}</b>
                                    &nbsp;&nbsp;
                                    <span style='color:{diff_col};font-weight:700;'>
                                        ({diff_sign}${m['diff']:.2f})
                                    </span>
                                </div>""", unsafe_allow_html=True)
                    else:
                        st.markdown("""
                        <div class='match-row'>
                            <b>✓ All rates match your rate table</b>
                        </div>""", unsafe_allow_html=True)

                    # ── Download button ──
                    excel_bytes = build_excel(results, all_mismatches)
                    fname = f"GrainCorp_Swaps_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                    st.download_button(
                        label="📥  Download Excel File",
                        data=excel_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )

                    # ── Data preview ──
                    with st.expander("👁️ Preview extracted data", expanded=False):
                        for r in results:
                            swap_no = str(r.get("swap_no",""))
                            mm      = all_mismatches.get(swap_no, [])
                            mm_fields = {m["field"] for m in mm}

                            cols = st.columns([2,2,2,2,1,1])
                            cols[0].metric("Swap No",  r.get("swap_no","—"))
                            cols[1].metric("From",     r.get("from_site_name","—"))
                            cols[2].metric("To",       r.get("to_site_name","—"))
                            cols[3].metric("Grade",    r.get("grade","—"))
                            cols[4].metric("Tonnes",   r.get("tonnes","—"))
                            cols[5].metric("Date",     r.get("transfer_date","—"))

                            r1, r2 = st.columns(2)
                            rf_flag = " ⚠️" if "Rate From" in mm_fields else " ✓"
                            rt_flag = " ⚠️" if "Rate To"   in mm_fields else " ✓"
                            r1.metric(f"Rate From{rf_flag}", f"${r.get('rate_from','—')}")
                            r2.metric(f"Rate To{rt_flag}",   f"${r.get('rate_to','—')}")
                            st.divider()
                else:
                    status.error("❌ No records extracted. Check your API key and try again.")
    else:
        st.markdown("""
        <div style='background:#F0F8F3;border:2px dashed #B8DCC8;border-radius:10px;
                    padding:40px;text-align:center;color:#6B7C72;'>
            <div style='font-size:36px;margin-bottom:8px;'>📄</div>
            <div style='font-size:16px;font-weight:600;color:#1B5E3B;'>
                Upload your Swap PDFs above</div>
            <div style='font-size:13px;margin-top:4px;'>
                Supports multiple files at once</div>
        </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
# HISTORY TAB
# ══════════════════════════════════════════════════════════════════════════
with tab_history:
    h = st.session_state.history
    col1, col2 = st.columns([3,1])
    with col1:
        st.markdown(f"### Export History  —  {len(h)} record(s)")
    with col2:
        if h:
            buf = io.StringIO()
            keys = ["swap_no","transfer_date","from_site_name","to_site_name",
                    "grade","season","tonnes","rate_from","rate_to",
                    "mismatches","source_file","extracted_at"]
            w = csv.DictWriter(buf, fieldnames=keys, extrasaction="ignore")
            w.writeheader(); w.writerows(h)
            st.download_button("📥 Export CSV", buf.getvalue(),
                               "GrainCorp_History.csv", "text/csv")

    if not h:
        st.info("No history yet — convert some PDFs to see them here.")
    else:
        rows_html = ""
        for entry in h:
            mm_count = entry.get("mismatches", 0)
            mm_badge = (f"<span class='badge-warn'>⚠ {mm_count} mismatch(es)</span>"
                        if mm_count else "<span class='badge-ok'>✓ OK</span>")
            rows_html += f"""<tr>
                <td><code>{entry.get('swap_no','—')}</code></td>
                <td>{entry.get('transfer_date','—')}</td>
                <td>{entry.get('from_site_name','—')}</td>
                <td>{entry.get('to_site_name','—')}</td>
                <td><b>{entry.get('grade','—')}</b></td>
                <td>{entry.get('tonnes','—')}</td>
                <td>${entry.get('rate_from','—')}</td>
                <td>${entry.get('rate_to','—')}</td>
                <td>{mm_badge}</td>
                <td style='color:#9DAAA2;font-size:11px;'>{entry.get('extracted_at','—')}</td>
            </tr>"""

        st.markdown(f"""<table class='hist-table'>
            <thead><tr>
                <th>Swap No</th><th>Date</th><th>From Site</th><th>To Site</th>
                <th>Grade</th><th>Tonnes</th><th>Rate From</th><th>Rate To</th>
                <th>Rates</th><th>Extracted At</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
        </table>""", unsafe_allow_html=True)

        if st.button("🗑️ Clear History"):
            st.session_state.history = []
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════
# LOG TAB
# ══════════════════════════════════════════════════════════════════════════
with tab_log:
    col1, col2 = st.columns([3,1])
    with col1: st.markdown("### Activity Log")
    with col2:
        if st.button("🗑️ Clear Log"):
            st.session_state.log = []; st.rerun()

    if not st.session_state.log:
        st.info("No activity yet.")
    else:
        colours = {"ok":"#1B7A3E","error":"#C0392B","warn":"#B7600A","info":"#1565C0"}
        for entry in st.session_state.log:
            c = colours.get(entry["level"], "#2C3E30")
            st.markdown(f"""
            <div style='padding:8px 12px;border-bottom:1px solid #F0F1F2;
                        font-family:Consolas,monospace;font-size:13px;'>
                <span style='color:#9DAAA2;'>[{entry["ts"]}]</span>
                <span style='color:{c};margin-left:8px;'>{entry["msg"]}</span>
            </div>""", unsafe_allow_html=True)
